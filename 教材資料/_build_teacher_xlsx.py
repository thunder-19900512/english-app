# -*- coding: utf-8 -*-
import re, os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
OUT = os.path.join(BASE, '教材資料', 'EnglishApp_教員用_単語文章一覧.xlsx')

# ---- 語彙を vocabulary.ts から抽出 ----
voc_src = open(os.path.join(BASE, 'src/data/vocabulary.ts'), encoding='utf-8').read()
pat = re.compile(r"english:\s*'([^']*)',\s*japanese:\s*'([^']*)',\s*category:\s*'([^']*)',\s*page:\s*(\d+),\s*emoji:\s*'[^']*',\s*keyPhrase:\s*\"([^\"]*)\"")
vocab = [{'english': m[0], 'japanese': m[1], 'category': m[2], 'page': int(m[3]), 'keyPhrase': m[4]} for m in pat.findall(voc_src)]

# ---- スタイル ----
HEAD_FILL = PatternFill('solid', start_color='4F81BD')
HEAD_FONT = Font(bold=True, color='FFFFFF', size=11)
TITLE_FONT = Font(bold=True, size=14, color='1F4E78')
WRAP = Alignment(wrap_text=True, vertical='top')
CENTER = Alignment(horizontal='center', vertical='center')
thin = Side(style='thin', color='BBBBBB')
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

def style_header(ws, row=1):
    for c in ws[row]:
        c.fill = HEAD_FILL; c.font = HEAD_FONT; c.alignment = CENTER; c.border = BORDER

def add_borders(ws, start=2):
    for row in ws.iter_rows(min_row=start):
        for c in row:
            c.border = BORDER
            if c.alignment.wrap_text is not True:
                c.alignment = Alignment(vertical='top')

wb = Workbook()

# ===== シート0: 概要 =====
ws = wb.active; ws.title = '概要'
ws['A1'] = 'EnglishApp 教員用 単語・文章一覧'; ws['A1'].font = TITLE_FONT
ws['A2'] = '※ アプリ内で実際に使われている語彙・文章のまとめです。「おはなしづくり」「AI英会話」はAIがその場で生成するため固定リストはありません。'
ws['A2'].alignment = WRAP
rows = [
    ['', ''],
    ['シート', '内容'],
    ['語彙一覧', 'Picture Dictionary（絵辞典）の全単語。学習/選択/タイピング/言葉さがし/モンスターバトル/QAで使用'],
    ['キーフレーズ早見表', 'カテゴリごとの基本文（型）。QAモード・各種練習で使う言い回し'],
    ['QAモードの質問', 'QAモードでアプリが出す質問文の一覧'],
    ['教科書モード', '学年×Unitのクイズ（問題・答え）とキー表現'],
    ['ダイアログ', 'ダイアログ・トレーナーの会話文（ペアワーク事前練習）'],
    ['フォニックス', 'Phonics 各ステージで扱う文字・単語・文'],
]
for r in rows: ws.append(r)
ws['A4'].fill = HEAD_FILL; ws['B4'].fill = HEAD_FILL
ws['A4'].font = HEAD_FONT; ws['B4'].font = HEAD_FONT
ws.column_dimensions['A'].width = 22; ws.column_dimensions['B'].width = 80
for row in ws.iter_rows(min_row=5):
    for c in row: c.alignment = WRAP

# ===== シート1: 語彙一覧 =====
ws = wb.create_sheet('語彙一覧')
ws.append(['カテゴリ', '英語', '日本語', 'キーフレーズ（型）', '教科書ページ'])
for v in vocab:
    ws.append([v['category'], v['english'], v['japanese'], v['keyPhrase'], v['page']])
widths = [16, 18, 20, 22, 12]
for i, w in enumerate(widths, 1): ws.column_dimensions[chr(64+i)].width = w
style_header(ws); add_borders(ws)
ws.freeze_panes = 'A2'; ws.auto_filter.ref = f'A1:E{ws.max_row}'

# ===== シート2: キーフレーズ早見表 =====
ws = wb.create_sheet('キーフレーズ早見表')
ws.append(['カテゴリ', 'キーフレーズ（型）', '語数'])
seen = {}
for v in vocab:
    seen.setdefault(v['category'], {'kp': v['keyPhrase'], 'n': 0})
    seen[v['category']]['n'] += 1
for cat, d in seen.items():
    ws.append([cat, d['kp'], d['n']])
for i, w in enumerate([18, 28, 8], 1): ws.column_dimensions[chr(64+i)].width = w
style_header(ws); add_borders(ws)

# ===== シート3: QAモードの質問 =====
ws = wb.create_sheet('QAモードの質問')
qa = [
    ("I'd like ◯◯. (食べ物/飲み物/デザート)", 'What would you like?'),
    ('One ◯◯, please. (果物・野菜)', 'May I take your order?'),
    ("Let's eat ◯◯. (食事)", 'What shall we eat?'),
    ('◯◯, please. (食材)', 'What would you like?'),
    ("It's ◯◯. (味など/曜日)", 'How is it?'),
    ('I can ◯◯ well. (動作5年)', 'What can you do?'),
    ('I want to ◯◯. (動作6年)', 'What do you want to do?'),
    ('I ◯◯ 〜. (したこと)', 'What did you do?'),
    ("I'm ◯◯ years old. (数)", 'How old are you?'),
    ("I'm ◯◯. (気分)", 'How are you?'),
    ('I like ◯◯. (色/形/季節)', 'What do you like?'),
    ('My birthday is in ◯◯. (月)', 'When is your birthday?'),
]
ws.append(['答えの型（キーフレーズ）', 'アプリが出す質問'])
for a, q in qa: ws.append([a, q])
for i, w in enumerate([34, 34], 1): ws.column_dimensions[chr(64+i)].width = w
style_header(ws); add_borders(ws)

# ===== シート4: 教科書モード =====
textbook = [
    ('5年', 'Unit 1: Hello, friends!', 'What subject do you like?（何の教科が好き？）', [
        ('ソフィアはどこの国から来た？', 'Australia'),
        ('ソフィアが好きな教科は？', 'P.E.'),
        ('ソフィアが紹介したオーストラリアの食べ物は？', 'Crocodile steak'),
    ]),
    ('5年', 'Unit 2: Happy birthday!', 'When is your birthday?（誕生日はいつ？）', [
        ('ルーカスが誕生日にほしいものは？', 'A new tablet'),
        ('みんながソフィアにあげたものは？', 'rugby sticker'),
        ('ソフィアの母が作ったステーキの肉は？', 'Beef'),
    ]),
    ('5年', 'Unit 3: What do you have on Monday?', 'Can you play the piano?（ピアノを弾ける？）', [
        ('子どもたちが楽しんでいた遊びは？', 'Dodgeball'),
        ('リコーダーできない人はピアノを弾ける？', 'Yes, I can.'),
        ('廊下を走る2人に先生は何と注意した？', "Don't run"),
    ]),
    ('5年', 'Unit 4: He can bake bread well.', 'She can play tennis very well.（彼女はテニスが上手）', [
        ('ジェシカが得意なスポーツは？', 'Tennis'),
        ('動物園で抱っこできる動物は？', 'koala'),
        ('「シュークリーム」は英語で？', 'Cream puffs'),
    ]),
    ('5年', 'Unit 5: Where is the post office?', 'Where is the post office?（郵便局はどこ？）', [
        ('郵便局へは何ブロックまっすぐ？', 'Two blocks'),
        ('動物園の看板の近くにいた動物は？', 'Cats'),
        ('ソフィアが電話で話していた相手は？', 'Dad'),
    ]),
    ('5年', 'Unit 6: What would you like?', 'What would you like?（何になさいますか？）', [
        ('ソフィアが注文したものは？', 'Beef bowl'),
        ('おばあちゃんが頼んだ飲み物は？', 'Orange juice'),
        ('オリバーが食べた焼きそばの味は？', 'A little spicy'),
    ]),
    ('5年', 'Unit 7: Welcome to Japan.', 'Why do you want to go there?（なぜそこへ行きたい？）', [
        ('弘前で春に見られるお祭りは？', 'Cherry Blossom Festival'),
        ('豊似湖はどんな形？', 'heart'),
        ('ソフィアが白川郷で食べたいものは？', 'Hida Beef'),
    ]),
    ('5年', 'Unit 8: Who is your hero?', 'Who is your hero?（あなたのヒーローは？）', [
        ('熱を出した子に母が渡したものは？', 'Medicine'),
        ('ベーカー先生のヒーロー角野栄子さんの職業は？', 'writer'),
        ('ソフィアのヒーローは？', 'Mother'),
    ]),
    ('6年', 'Unit 1: This is me!', 'I can speak Swahili and English.（スワヒリ語と英語が話せる）', [
        ('ナディアの出身国は？', 'Kenya'),
        ('ナディアの宝物は？', 'A sweatshirt'),
        ('ナディアの家は何の近く？', 'library'),
    ]),
    ('6年', 'Unit 2: How is your school life?', 'What time do you get up?（何時に起きる？）', [
        ('ヘルミが起きる時間は？', '6:00 a.m.'),
        ('NZの10:30のおやつの時間の呼び名は？', 'morning tea'),
        ('ダニエルは何を通って通学する？', 'The savanna'),
    ]),
    ('6年', 'Unit 3: My Summer Vacation', 'How was your weekend?（週末はどうだった？）', [
        ('サキが食べたのはどこの国の料理？', 'Swiss food'),
        ('ソフィアが楽しんだ試合のスポーツは？', 'rugby'),
        ('オールブラックスが踊るダンスの名前は？', 'Haka'),
    ]),
    ('6年', "Unit 4: Let's see the world.", 'You can see many unique animals.（めずらしい動物が見られる）', [
        ('オーストラリアの大きな岩の名前は？', 'Uluru'),
        ('ベトナムの伝統的なドレスの名前は？', 'ao dai'),
        ('熱帯雨林で有名と紹介された国は？', 'Brazil'),
    ]),
    ('6年', 'Unit 5: We live in a global village.', 'This sweater is from New Zealand.（このセーターはNZ産）', [
        ('帽子とセーターはどこの国のウール？', 'New Zealand'),
        ('タコはどこの国から来ている？', 'Morocco'),
        ('ピクニックで食べたいサンドイッチは？', 'BLT'),
    ]),
    ('6年', "Unit 6: Let's think about our food.", 'What do sea turtles eat?（ウミガメは何を食べる？）', [
        ('ウミガメが間違えて食べる海のゴミは？', 'Plastic bags'),
        ('マイバッグで減らせるものは？', 'Plastic'),
        ('物を再利用することを英語で？', 'reuse'),
    ]),
    ('6年', 'Unit 7: My Best Memory', 'My best memory is the school trip.（一番の思い出は修学旅行）', [
        ('修学旅行は英語で？', 'The school trip'),
        ('日光で猿はどこにいた？', 'On the roof'),
        ('美術部は英語で？', 'art club'),
    ]),
    ('6年', 'Unit 8: Future Dreams', 'I want to be a programmer.（プログラマーになりたい）', [
        ('ダイチの将来の夢は？', 'Programmer'),
        ('ナディアが将来なりたいものは？', 'Vet (獣医)'),
        ('ルーカスが将来住みたい国は？', 'Japan'),
    ]),
]
ws = wb.create_sheet('教科書モード')
ws.append(['学年', 'Unit', 'キー表現', '問題', '答え'])
for grade, unit, kp, qs in textbook:
    for i, (q, a) in enumerate(qs):
        ws.append([grade if i == 0 else '', unit if i == 0 else '', kp if i == 0 else '', q, a])
for i, w in enumerate([6, 30, 34, 40, 22], 1): ws.column_dimensions[chr(64+i)].width = w
style_header(ws); add_borders(ws)
for row in ws.iter_rows(min_row=2):
    row[2].alignment = WRAP; row[3].alignment = WRAP

# ===== シート5: ダイアログ =====
dialogues = [
    ('5年', 'Unit 1', 'What subject do you like?', [
        ('A', 'What subject do you like?', '何の教科が好き？'),
        ('B', 'I like {P.E.}.', '【体育】が好き。'),
        ('A', 'Nice! I like {music}, too.', 'いいね！ぼくは【音楽】が好き。')]),
    ('5年', 'Unit 2', 'When is your birthday?', [
        ('A', 'When is your birthday?', '誕生日はいつ？'),
        ('B', 'My birthday is {May 5th}.', '【5月5日】だよ。'),
        ('A', 'Happy birthday! What do you want?', 'おめでとう！何がほしい？'),
        ('B', 'I want {a new ball}.', '【新しいボール】がほしい。')]),
    ('5年', 'Unit 3', 'Can you play the piano?', [
        ('A', 'Can you play the piano?', 'ピアノ弾ける？'),
        ('B', '{Yes, I can.}', '【うん、できるよ。】'),
        ('A', 'Can you {swim}?', '【泳ぐ】のはできる？'),
        ('B', 'Yes, I can!', 'うん、できる！')]),
    ('5年', 'Unit 4', 'She can play tennis very well.', [
        ('A', 'This is my friend, {Ken}.', '友だちの【ケン】だよ。'),
        ('B', 'What can {he} do?', '【彼】は何ができるの？'),
        ('A', '{He} can play tennis very well.', '【彼】はテニスがすごく上手。'),
        ('B', 'Wow, great!', 'わー、すごい！')]),
    ('5年', 'Unit 5', 'Where is the post office?', [
        ('A', 'Excuse me. Where is the {post office}?', 'すみません、【郵便局】はどこ？'),
        ('B', 'Go straight and turn {right}.', 'まっすぐ行って【右】へ。'),
        ('A', 'Thank you!', 'ありがとう！')]),
    ('5年', 'Unit 6', 'What would you like?', [
        ('A', 'What would you like?', '何にする？'),
        ('B', "I'd like {a hamburger}.", '【ハンバーガー】をください。'),
        ('A', 'Anything to drink?', '飲み物は？'),
        ('B', '{Orange juice}, please.', '【オレンジジュース】を。')]),
    ('5年', 'Unit 7', 'Why do you want to go there?', [
        ('A', 'Where do you want to go?', 'どこに行きたい？'),
        ('B', 'I want to go to {Okinawa}.', '【沖縄】に行きたい。'),
        ('A', 'Why do you want to go there?', 'なんで？'),
        ('B', 'I want to {see the sea}.', '【海が見たい】から。')]),
    ('5年', 'Unit 8', 'Who is your hero?', [
        ('A', 'Who is your hero?', 'ヒーローは誰？'),
        ('B', 'My hero is my {mother}.', '【お母さん】だよ。'),
        ('A', 'Why?', 'どうして？'),
        ('B', '{She} is {kind}.', '【やさしい】から。')]),
    ('6年', 'Unit 1', 'I can speak Swahili and English.', [
        ('A', "Hi, I'm {Sora}. Nice to meet you.", '【そら】だよ、よろしく。'),
        ('B', 'Nice to meet you, too. What can you do?', 'よろしく。何ができる？'),
        ('A', 'I can speak {Japanese} and {English}.', '【日本語】と【英語】が話せる。')]),
    ('6年', 'Unit 2', 'What time do you get up?', [
        ('A', 'What time do you get up?', '何時に起きる？'),
        ('B', 'I get up at {6:30}.', '【6時半】に起きる。'),
        ('A', 'What time do you go to bed?', '何時に寝る？'),
        ('B', 'I go to bed at {9:30}.', '【9時半】に寝る。')]),
    ('6年', 'Unit 3', 'How was your weekend?', [
        ('A', 'How was your weekend?', '週末どうだった？'),
        ('B', 'It was {fun}. I {played soccer}.', '【楽しかった】。【サッカーした】。'),
        ('A', 'Sounds nice!', 'いいね！')]),
    ('6年', 'Unit 4', 'You can see many unique animals.', [
        ('A', 'Where should we go?', 'どこ行く？'),
        ('B', "Let's go to {Australia}.", '【オーストラリア】に行こう。'),
        ('A', 'Why?', 'なんで？'),
        ('B', 'You can see many {unique animals}.', '【めずらしい動物】がたくさん見られる。')]),
    ('6年', 'Unit 5', 'This sweater is from New Zealand.', [
        ('A', 'Nice {sweater}! Where is it from?', 'いい【セーター】！どこ産？'),
        ('B', 'This {sweater} is from {New Zealand}.', '【ニュージーランド】産だよ。'),
        ('A', "Wow, that's far!", '遠いね！')]),
    ('6年', 'Unit 6', 'What do sea turtles eat?', [
        ('A', '{Sea turtles} are in danger.', '【ウミガメ】が危ない。'),
        ('B', 'They live in {the sea}.', '【海】に住んでいる。'),
        ('A', 'We must reduce {plastic bags}.', '【レジ袋】を減らさなきゃ。')]),
    ('6年', 'Unit 7', 'My best memory is the school trip.', [
        ('A', 'What is your best memory?', '一番の思い出は？'),
        ('B', 'My best memory is {the school trip}.', '【修学旅行】。'),
        ('A', 'What did you do?', '何をした？'),
        ('B', 'I {saw a big castle}.', '【大きなお城を見た】。')]),
    ('6年', 'Unit 8', 'I want to be a programmer.', [
        ('A', 'What do you want to be?', '将来何になりたい？'),
        ('B', 'I want to be {a vet}.', '【獣医】になりたい。'),
        ('A', 'Why?', 'どうして？'),
        ('B', 'I like {animals}.', '【動物】が好きだから。')]),
]
ws = wb.create_sheet('ダイアログ')
ws.append(['学年', 'Unit', 'キー表現', '役', '英語（{ }は差し替え）', '日本語'])
for grade, unit, kp, lines in dialogues:
    for i, (sp, en, ja) in enumerate(lines):
        ws.append([grade if i == 0 else '', unit if i == 0 else '', kp if i == 0 else '', sp, en, ja])
for i, w in enumerate([6, 9, 30, 6, 40, 30], 1): ws.column_dimensions[chr(64+i)].width = w
style_header(ws); add_borders(ws)
for row in ws.iter_rows(min_row=2):
    row[2].alignment = WRAP; row[4].alignment = WRAP; row[5].alignment = WRAP

# ===== シート6: フォニックス =====
phonics = [
    (1, 'アルファベットの音', 'A B C D E F G H I J K L M N O P Q R S T U V W X Y Z', '', 'bif, mup, zat', ''),
    (2, '3文字の単語', 'cat dog pig bed sun hat bat map pin lip box fox cup bug net red', '', 'zop, lat, mab, ren, gug',
     'The cat is on the bed. / A bug is in the cup. / The pig has a map. / The dog has a red hat.'),
    (3, '2文字で1つの音', 'ship shop fish dish star stop chop chin thin math phone whale duck ring queen', 'sh ch th ph wh ck ng',
     'shup, chog, thip, whet, fick', 'The fish is in the shop. / A duck is on the ship. / I see a ring and a dish.'),
    (4, 'マジック E', 'cake make take bike like kite nose rose cute mute', 'cap→cape / kit→kite / cub→cube / pin→pine',
     'vane, fite, moke, zute, bape', 'I like the cute rose. / Take the bike and make a cake. / His nose is on the kite.'),
    (5, '母音のペア', 'rain train tree see boat coat meat seat', 'ai ea ee oa', 'zaip, meeb, foat, steab, traif',
     'I see a tree in the rain. / The meat is on the boat. / Take a seat on the train.'),
    (6, '混ざり合った音', 'smile snack black clock spoon stop park short girl chair work', 'ar or ir air ear wor',
     'darf, lorn, mirt, slear, worf', 'The girl is in the park. / I see a short clock on the chair. / A black dog has a snack.'),
]
ws = wb.create_sheet('フォニックス')
ws.append(['ステージ', 'テーマ', '練習する文字', '単語', 'エイリアン語(無意味語)', '文（音読）'])
for sid, theme, items, pract, alien, stories in phonics:
    ws.append([f'Stage {sid}', theme, pract, items, alien, stories])
for i, w in enumerate([9, 18, 22, 40, 24, 44], 1): ws.column_dimensions[chr(64+i)].width = w
style_header(ws); add_borders(ws)
for row in ws.iter_rows(min_row=2):
    for c in row: c.alignment = WRAP

wb.save(OUT)
print('SAVED', OUT)
print('vocab count:', len(vocab))
